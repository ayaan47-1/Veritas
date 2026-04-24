from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterable, Iterator
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Literal
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth.deps import get_current_user, require_asset_scope
from ..config import settings
from ..database import get_db
from ..models import (
    Asset,
    Document,
    Obligation,
    ObligationEvidence,
    ObligationReview,
    ReviewStatus,
    Risk,
    RiskEvidence,
    RiskReview,
    RiskType,
    Severity,
    User,
)

router = APIRouter(prefix="/exports", tags=["exports"])


OBLIGATION_COLUMNS: list[str] = [
    "id",
    "asset_name",
    "document_filename",
    "obligation_type",
    "text",
    "severity",
    "system_confidence",
    "llm_quality_confidence",
    "status",
    "deadline",
    "evidence_quote",
    "evidence_page_number",
    "evidence_char_start",
    "evidence_char_end",
    "created_at",
    "last_reviewed_at",
    "reviewer_email",
]


RISK_COLUMNS: list[str] = [
    "id",
    "asset_name",
    "document_filename",
    "risk_type",
    "text",
    "severity",
    "system_confidence",
    "llm_quality_confidence",
    "status",
    "evidence_quote",
    "evidence_page_number",
    "evidence_char_start",
    "evidence_char_end",
    "created_at",
    "last_reviewed_at",
    "reviewer_email",
]


_SEVERITY_FILL_HEX: dict[str, str] = {
    "critical": "FFEF4444",
    "high": "FFF97316",
    "medium": "FFEAB308",
    "low": "FF3B82F6",
}


def _slug(name: str | None) -> str:
    if not name:
        return "all"
    lowered = name.lower()
    replaced = re.sub(r"\s+", "_", lowered)
    stripped = re.sub(r"[^a-z0-9_-]", "", replaced)
    collapsed = re.sub(r"_+", "_", stripped).strip("_-")
    return collapsed or "all"


def _filename(entity: str, asset_name: str | None, ext: str) -> str:
    return f"{entity}_{_slug(asset_name)}_{datetime.now(tz=timezone.utc).date().isoformat()}.{ext}"


def _max_rows() -> int:
    raw = settings.raw.get("exports", {}).get("max_rows", 50000)
    return int(raw)


def _build_obligation_query(
    db: Session,
    *,
    status: ReviewStatus | None,
    severity: Severity | None,
    document_id: UUID | None,
    asset_id: UUID | None,
):
    query = db.query(Obligation)
    if status is not None:
        query = query.filter(Obligation.status == status)
    if severity is not None:
        query = query.filter(Obligation.severity == severity)
    if document_id is not None:
        query = query.filter(Obligation.document_id == document_id)
    if asset_id is not None:
        query = query.join(Document, Obligation.document_id == Document.id).filter(Document.asset_id == asset_id)
    return query


def _build_risk_query(
    db: Session,
    *,
    status: ReviewStatus | None,
    severity: Severity | None,
    risk_type: RiskType | None,
    document_id: UUID | None,
    asset_id: UUID | None,
):
    query = db.query(Risk)
    if status is not None:
        query = query.filter(Risk.status == status)
    if severity is not None:
        query = query.filter(Risk.severity == severity)
    if risk_type is not None:
        query = query.filter(Risk.risk_type == risk_type)
    if document_id is not None:
        query = query.filter(Risk.document_id == document_id)
    if asset_id is not None:
        query = query.join(Document, Risk.document_id == Document.id).filter(Document.asset_id == asset_id)
    return query


@dataclass
class _ObligationRow:
    obligation: Obligation
    asset_name: str
    document_filename: str
    evidence: ObligationEvidence | None
    last_review: ObligationReview | None
    reviewer_email: str | None


@dataclass
class _RiskRow:
    risk: Risk
    asset_name: str
    document_filename: str
    evidence: RiskEvidence | None
    last_review: RiskReview | None
    reviewer_email: str | None


def _iso_or_empty(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _severity_for_export(system: Severity, llm: Severity | None) -> str:
    return (llm.value if llm is not None else system.value)


def _resolve_obligation_rows(
    db: Session, obligations: list[Obligation]
) -> Iterator[_ObligationRow]:
    if not obligations:
        return iter(())

    obligation_ids = [row.id for row in obligations]
    document_ids = list({row.document_id for row in obligations})

    documents = {d.id: d for d in db.query(Document).filter(Document.id.in_(document_ids)).all()}
    asset_ids = list({d.asset_id for d in documents.values()})
    assets = {a.id: a for a in db.query(Asset).filter(Asset.id.in_(asset_ids)).all()} if asset_ids else {}

    evidence_rows = (
        db.query(ObligationEvidence)
        .filter(ObligationEvidence.obligation_id.in_(obligation_ids))
        .order_by(ObligationEvidence.created_at.asc(), ObligationEvidence.id.asc())
        .all()
    )
    primary_evidence: dict[UUID, ObligationEvidence] = {}
    for ev in evidence_rows:
        primary_evidence.setdefault(ev.obligation_id, ev)

    review_rows = (
        db.query(ObligationReview)
        .filter(ObligationReview.obligation_id.in_(obligation_ids))
        .order_by(ObligationReview.created_at.desc(), ObligationReview.id.desc())
        .all()
    )
    last_review: dict[UUID, ObligationReview] = {}
    for rv in review_rows:
        last_review.setdefault(rv.obligation_id, rv)

    reviewer_ids = list({rv.reviewer_id for rv in last_review.values() if rv.reviewer_id is not None})
    reviewer_emails = (
        {u.id: u.email for u in db.query(User).filter(User.id.in_(reviewer_ids)).all()}
        if reviewer_ids
        else {}
    )

    def _iter() -> Iterator[_ObligationRow]:
        for ob in obligations:
            document = documents.get(ob.document_id)
            asset = assets.get(document.asset_id) if document else None
            rv = last_review.get(ob.id)
            yield _ObligationRow(
                obligation=ob,
                asset_name=asset.name if asset else "",
                document_filename=document.source_name if document else "",
                evidence=primary_evidence.get(ob.id),
                last_review=rv,
                reviewer_email=reviewer_emails.get(rv.reviewer_id) if rv and rv.reviewer_id else None,
            )

    return _iter()


def _resolve_risk_rows(db: Session, risks: list[Risk]) -> Iterator[_RiskRow]:
    if not risks:
        return iter(())

    risk_ids = [row.id for row in risks]
    document_ids = list({row.document_id for row in risks})

    documents = {d.id: d for d in db.query(Document).filter(Document.id.in_(document_ids)).all()}
    asset_ids = list({d.asset_id for d in documents.values()})
    assets = {a.id: a for a in db.query(Asset).filter(Asset.id.in_(asset_ids)).all()} if asset_ids else {}

    evidence_rows = (
        db.query(RiskEvidence)
        .filter(RiskEvidence.risk_id.in_(risk_ids))
        .order_by(RiskEvidence.created_at.asc(), RiskEvidence.id.asc())
        .all()
    )
    primary_evidence: dict[UUID, RiskEvidence] = {}
    for ev in evidence_rows:
        primary_evidence.setdefault(ev.risk_id, ev)

    review_rows = (
        db.query(RiskReview)
        .filter(RiskReview.risk_id.in_(risk_ids))
        .order_by(RiskReview.created_at.desc(), RiskReview.id.desc())
        .all()
    )
    last_review: dict[UUID, RiskReview] = {}
    for rv in review_rows:
        last_review.setdefault(rv.risk_id, rv)

    reviewer_ids = list({rv.reviewer_id for rv in last_review.values() if rv.reviewer_id is not None})
    reviewer_emails = (
        {u.id: u.email for u in db.query(User).filter(User.id.in_(reviewer_ids)).all()}
        if reviewer_ids
        else {}
    )

    def _iter() -> Iterator[_RiskRow]:
        for r in risks:
            document = documents.get(r.document_id)
            asset = assets.get(document.asset_id) if document else None
            rv = last_review.get(r.id)
            yield _RiskRow(
                risk=r,
                asset_name=asset.name if asset else "",
                document_filename=document.source_name if document else "",
                evidence=primary_evidence.get(r.id),
                last_review=rv,
                reviewer_email=reviewer_emails.get(rv.reviewer_id) if rv and rv.reviewer_id else None,
            )

    return _iter()


def _row_for_obligation(row: _ObligationRow) -> list[str]:
    ob = row.obligation
    ev = row.evidence
    last = row.last_review
    return [
        str(ob.id),
        row.asset_name,
        row.document_filename,
        ob.obligation_type.value,
        ob.obligation_text or "",
        _severity_for_export(ob.severity, ob.llm_severity),
        _iso_or_empty(ob.system_confidence),
        _iso_or_empty(ob.llm_quality_confidence),
        ob.status.value,
        _iso_or_empty(ob.due_date),
        (ev.quote if ev else ""),
        _iso_or_empty(ev.page_number if ev else None),
        _iso_or_empty(ev.raw_char_start if ev else None),
        _iso_or_empty(ev.raw_char_end if ev else None),
        _iso_or_empty(ob.created_at),
        _iso_or_empty(last.created_at if last else None),
        row.reviewer_email or "",
    ]


def _row_for_risk(row: _RiskRow) -> list[str]:
    r = row.risk
    ev = row.evidence
    last = row.last_review
    return [
        str(r.id),
        row.asset_name,
        row.document_filename,
        r.risk_type.value,
        r.risk_text or "",
        _severity_for_export(r.severity, r.llm_severity),
        _iso_or_empty(r.system_confidence),
        _iso_or_empty(r.llm_quality_confidence),
        r.status.value,
        (ev.quote if ev else ""),
        _iso_or_empty(ev.page_number if ev else None),
        _iso_or_empty(ev.raw_char_start if ev else None),
        _iso_or_empty(ev.raw_char_end if ev else None),
        _iso_or_empty(r.created_at),
        _iso_or_empty(last.created_at if last else None),
        row.reviewer_email or "",
    ]


def _csv_lines(columns: list[str], rows: Iterable[list[str]]) -> Iterator[str]:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate()
    for row in rows:
        writer.writerow(row)
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate()


def _xlsx_bytes(
    columns: list[str],
    rows: Iterable[list[str]],
    sheet_name: str,
    severity_column_index: int,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    widths = {
        "id": 38,
        "asset_name": 24,
        "document_filename": 30,
        "obligation_type": 16,
        "risk_type": 16,
        "text": 60,
        "severity": 12,
        "system_confidence": 12,
        "llm_quality_confidence": 14,
        "status": 14,
        "deadline": 14,
        "evidence_quote": 60,
        "evidence_page_number": 10,
        "evidence_char_start": 12,
        "evidence_char_end": 12,
        "created_at": 24,
        "last_reviewed_at": 24,
        "reviewer_email": 28,
    }
    for idx, column_name in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = widths.get(column_name, 18)

    severity_column_excel_index = severity_column_index + 1  # 1-based
    for row in rows:
        sheet.append(row)
        sev_value = row[severity_column_index]
        hex_fill = _SEVERITY_FILL_HEX.get(sev_value)
        if hex_fill:
            cell = sheet.cell(row=sheet.max_row, column=severity_column_excel_index)
            cell.fill = PatternFill(start_color=hex_fill, end_color=hex_fill, fill_type="solid")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _too_large(count: int) -> HTTPException | None:
    cap = _max_rows()
    if count > cap:
        return HTTPException(
            status_code=413,
            detail=f"Export exceeds {cap} rows; tighten filters",
        )
    return None


@router.get(
    "/obligations",
    dependencies=[Depends(require_asset_scope("asset_id", required_for_non_admin=True))],
)
def export_obligations(
    format: Literal["csv", "xlsx"] = Query(default="csv"),
    status: ReviewStatus | None = Query(default=None),
    severity: Severity | None = Query(default=None),
    document_id: UUID | None = Query(default=None),
    asset_id: UUID | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = _build_obligation_query(
        db,
        status=status,
        severity=severity,
        document_id=document_id,
        asset_id=asset_id,
    )
    count = query.count()
    error = _too_large(count)
    if error is not None:
        raise error

    obligations = query.order_by(Obligation.created_at.desc()).all()
    rows = list(_resolve_obligation_rows(db, obligations))
    asset_name = rows[0].asset_name if rows and asset_id is not None else None
    filename = _filename("obligations", asset_name, "xlsx" if format == "xlsx" else "csv")

    severity_index = OBLIGATION_COLUMNS.index("severity")

    if format == "xlsx":
        data = _xlsx_bytes(
            OBLIGATION_COLUMNS,
            (_row_for_obligation(row) for row in rows),
            sheet_name="Obligations",
            severity_column_index=severity_index,
        )
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    stream = _csv_lines(OBLIGATION_COLUMNS, (_row_for_obligation(row) for row in rows))
    return StreamingResponse(
        stream,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/risks",
    dependencies=[Depends(require_asset_scope("asset_id", required_for_non_admin=True))],
)
def export_risks(
    format: Literal["csv", "xlsx"] = Query(default="csv"),
    status: ReviewStatus | None = Query(default=None),
    severity: Severity | None = Query(default=None),
    risk_type: RiskType | None = Query(default=None),
    document_id: UUID | None = Query(default=None),
    asset_id: UUID | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = _build_risk_query(
        db,
        status=status,
        severity=severity,
        risk_type=risk_type,
        document_id=document_id,
        asset_id=asset_id,
    )
    count = query.count()
    error = _too_large(count)
    if error is not None:
        raise error

    risks = query.order_by(Risk.created_at.desc()).all()
    rows = list(_resolve_risk_rows(db, risks))
    asset_name = rows[0].asset_name if rows and asset_id is not None else None
    filename = _filename("risks", asset_name, "xlsx" if format == "xlsx" else "csv")

    severity_index = RISK_COLUMNS.index("severity")

    if format == "xlsx":
        data = _xlsx_bytes(
            RISK_COLUMNS,
            (_row_for_risk(row) for row in rows),
            sheet_name="Risks",
            severity_column_index=severity_index,
        )
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    stream = _csv_lines(RISK_COLUMNS, (_row_for_risk(row) for row in rows))
    return StreamingResponse(
        stream,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
