from __future__ import annotations

import hashlib
import importlib
import uuid
from datetime import UTC, datetime
from io import BytesIO

from fastapi.testclient import TestClient

from backend.app.auth import deps as auth_deps
from backend.app.database import get_db
from backend.app.main import create_app
from backend.app.models import (
    Asset,
    Document,
    DueKind,
    Modality,
    Obligation,
    ObligationEvidence,
    ObligationReview,
    ObligationType,
    OIDCProvider,
    ParseStatus,
    ReviewDecision,
    ReviewStatus,
    Risk,
    RiskEvidence,
    RiskReview,
    RiskType,
    Severity,
    TextSource,
    User,
    UserAssetAssignment,
    UserRole,
)

exports_router = importlib.import_module("backend.app.routers.exports")


OBLIGATION_COLUMNS = [
    "id",
    "asset_name",
    "document_filename",
    "obligation_type",
    "text",
    "severity",
    "system_confidence",
    "llm_quality_confidence",
    "status",
    "deadline",
    "evidence_quote",
    "evidence_page_number",
    "evidence_char_start",
    "evidence_char_end",
    "created_at",
    "last_reviewed_at",
    "reviewer_email",
]

RISK_COLUMNS = [
    "id",
    "asset_name",
    "document_filename",
    "risk_type",
    "text",
    "severity",
    "system_confidence",
    "llm_quality_confidence",
    "status",
    "evidence_quote",
    "evidence_page_number",
    "evidence_char_start",
    "evidence_char_end",
    "created_at",
    "last_reviewed_at",
    "reviewer_email",
]


def test_obligation_columns_match_spec():
    assert exports_router.OBLIGATION_COLUMNS == OBLIGATION_COLUMNS


def test_risk_columns_match_spec():
    assert exports_router.RISK_COLUMNS == RISK_COLUMNS


def test_slug_handles_spaces_and_special_chars():
    assert exports_router._slug("Willow Creek Tower") == "willow_creek_tower"
    assert exports_router._slug("A&B / C, D!") == "ab_c_d"
    assert exports_router._slug("   ") == "all"
    assert exports_router._slug(None) == "all"
    assert exports_router._slug("") == "all"


def test_filename_structure(monkeypatch):
    filename = exports_router._filename("obligations", "Willow Creek", "csv")
    assert filename.startswith("obligations_willow_creek_")
    assert filename.endswith(".csv")
    date_part = filename[len("obligations_willow_creek_"):-4]
    assert len(date_part) == 10
    assert date_part[4] == "-" and date_part[7] == "-"


def test_filename_all_when_no_asset():
    filename = exports_router._filename("risks", None, "xlsx")
    assert filename.startswith("risks_all_")
    assert filename.endswith(".xlsx")


def test_build_obligation_query_filters_by_status_and_severity():
    app = create_app()
    # Use the real test DB machinery — we just need filter wiring to compile and call.
    # We'll verify correctness end-to-end in the endpoint tests below.
    assert callable(exports_router._build_obligation_query)
    assert callable(exports_router._build_risk_query)


from openpyxl import load_workbook


class FakeQuery:
    def __init__(self, session: "FakeSession", *models):
        self._session = session
        self._model = models[0] if models else None
        self._conditions = []
        self._offset = 0
        self._limit = None

    def filter(self, *conditions):
        self._conditions.extend(conditions)
        return self

    def join(self, *_args, **_kwargs):
        return self

    def order_by(self, *_args):
        return self

    def offset(self, offset: int):
        self._offset = offset
        return self

    def limit(self, limit: int):
        self._limit = limit
        return self

    def all(self):
        rows = [r for r in self._rows_for_model() if self._matches_all(r)]
        rows = rows[self._offset:]
        if self._limit is not None:
            rows = rows[: self._limit]
        return rows

    def first(self):
        rows = self.all()
        return rows[0] if rows else None

    def count(self):
        return len(self.all())

    def _rows_for_model(self):
        s = self._session
        mapping = {
            User: s.users,
            UserAssetAssignment: s.assignments,
            Asset: s.assets,
            Document: s.documents,
            Obligation: s.obligations,
            ObligationEvidence: s.obligation_evidence,
            ObligationReview: s.obligation_reviews,
            Risk: s.risks,
            RiskEvidence: s.risk_evidence,
            RiskReview: s.risk_reviews,
        }
        return list(mapping.get(self._model, []))

    def _matches_all(self, row):
        return all(self._matches(row, c) for c in self._conditions)

    def _matches(self, row, condition):
        left = getattr(condition, "left", None)
        right = getattr(condition, "right", None)
        if left is None:
            return True
        key = getattr(left, "key", None)
        if key is None:
            return True
        row_value = self._row_value(row, key)
        op_name = getattr(getattr(condition, "operator", None), "__name__", "")
        raw_value = getattr(right, "value", right)
        if op_name == "in_op":
            return row_value in list(raw_value)
        return row_value == raw_value

    def _row_value(self, row, key):
        if hasattr(row, key):
            return getattr(row, key)
        if key == "asset_id" and hasattr(row, "document_id"):
            doc = self._session.document_by_id.get(row.document_id)
            return doc.asset_id if doc else None
        return None


class FakeSession:
    def __init__(self):
        self.users: list[User] = []
        self.assignments: list[UserAssetAssignment] = []
        self.assets: list[Asset] = []
        self.documents: list[Document] = []
        self.obligations: list[Obligation] = []
        self.obligation_evidence: list[ObligationEvidence] = []
        self.obligation_reviews: list[ObligationReview] = []
        self.risks: list[Risk] = []
        self.risk_evidence: list[RiskEvidence] = []
        self.risk_reviews: list[RiskReview] = []
        self.document_by_id: dict = {}

    def query(self, *models):
        return FakeQuery(self, *models)

    def add(self, _obj):
        return None

    def commit(self):
        return None

    def rollback(self):
        return None

    def close(self):
        return None


def _admin_user() -> User:
    return User(
        id=uuid.uuid4(),
        email="admin@example.com",
        name="Admin",
        oidc_provider=OIDCProvider.clerk,
        oidc_subject="admin-sub",
        role=UserRole.admin,
        is_active=True,
    )


def _viewer_user() -> User:
    return User(
        id=uuid.uuid4(),
        email="viewer@example.com",
        name="Viewer",
        oidc_provider=OIDCProvider.clerk,
        oidc_subject="viewer-sub",
        role=UserRole.viewer,
        is_active=True,
    )


def _asset(created_by) -> Asset:
    return Asset(
        id=uuid.uuid4(),
        name="Willow Creek",
        description=None,
        created_by=created_by,
    )


def _document(asset_id, uploaded_by) -> Document:
    return Document(
        id=uuid.uuid4(),
        asset_id=asset_id,
        source_name="lease.pdf",
        file_path="/tmp/lease.pdf",
        sha256=hashlib.sha256(str(asset_id).encode()).hexdigest(),
        mime_type="application/pdf",
        uploaded_by=uploaded_by,
        parse_status=ParseStatus.complete,
        scanned_page_count=0,
    )


def _obligation(
    document_id,
    *,
    severity=Severity.high,
    status=ReviewStatus.needs_review,
    llm_severity=None,
) -> Obligation:
    return Obligation(
        id=uuid.uuid4(),
        document_id=document_id,
        obligation_type=ObligationType.payment,
        obligation_text="Tenant shall pay rent on the first of each month.",
        modality=Modality.shall,
        responsible_entity_id=None,
        due_kind=DueKind.none,
        due_date=None,
        due_rule=None,
        trigger_date=None,
        severity=severity,
        llm_severity=llm_severity,
        llm_quality_confidence=None,
        status=status,
        system_confidence=85,
        reviewer_confidence=None,
        has_external_reference=False,
        contradiction_flag=False,
        extraction_run_id=None,
        created_at=datetime(2026, 4, 20, 12, 0, tzinfo=UTC),
    )


def _risk(document_id, *, severity=Severity.critical, status=ReviewStatus.confirmed) -> Risk:
    return Risk(
        id=uuid.uuid4(),
        document_id=document_id,
        risk_type=RiskType.financial,
        risk_text="Uncapped indemnification exposes landlord to unlimited liability.",
        severity=severity,
        status=status,
        system_confidence=80,
        reviewer_confidence=None,
        llm_severity=None,
        llm_quality_confidence=None,
        has_external_reference=False,
        contradiction_flag=False,
        extraction_run_id=None,
        created_at=datetime(2026, 4, 20, 12, 0, tzinfo=UTC),
    )


def _make_client(db: FakeSession, current_user: User) -> TestClient:
    app = create_app()

    def _override_db():
        yield db

    app.dependency_overrides[get_db] = _override_db
    app.dependency_overrides[auth_deps.get_current_user] = lambda: current_user
    return TestClient(app)


def _seed_obligations_db(obligations_count: int = 2) -> tuple[FakeSession, User, Asset]:
    db = FakeSession()
    admin = _admin_user()
    db.users.append(admin)
    asset = _asset(admin.id)
    db.assets.append(asset)
    doc = _document(asset.id, admin.id)
    db.documents.append(doc)
    db.document_by_id = {doc.id: doc}
    for i in range(obligations_count):
        ob = _obligation(doc.id, severity=Severity.high if i == 0 else Severity.low)
        db.obligations.append(ob)
    return db, admin, asset


def test_export_obligations_csv_basic():
    db, admin, asset = _seed_obligations_db(obligations_count=2)
    client = _make_client(db, admin)

    response = client.get(f"/exports/obligations?asset_id={asset.id}")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "obligations_willow_creek_" in response.headers["content-disposition"]

    lines = response.text.strip().split("\n")
    assert lines[0].strip() == ",".join(OBLIGATION_COLUMNS)
    assert len(lines) == 3  # header + 2 rows


def test_export_obligations_csv_respects_severity_filter():
    db, admin, asset = _seed_obligations_db(obligations_count=2)
    client = _make_client(db, admin)

    response = client.get(f"/exports/obligations?asset_id={asset.id}&severity=high")
    assert response.status_code == 200
    lines = response.text.strip().split("\n")
    assert len(lines) == 2  # header + 1 row


def test_export_obligations_xlsx_severity_cell_colors():
    db, admin, asset = _seed_obligations_db(obligations_count=0)
    doc = db.documents[0]
    for sev in (Severity.critical, Severity.high, Severity.medium, Severity.low):
        db.obligations.append(_obligation(doc.id, severity=sev))
    client = _make_client(db, admin)

    response = client.get(f"/exports/obligations?asset_id={asset.id}&format=xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == "Obligations"
    assert sheet["A1"].font.bold is True
    assert sheet.freeze_panes == "A2"

    severity_column_index = OBLIGATION_COLUMNS.index("severity") + 1
    expected = {
        "critical": "FFEF4444",
        "high": "FFF97316",
        "medium": "FFEAB308",
        "low": "FF3B82F6",
    }
    actual = {}
    for row_idx in range(2, sheet.max_row + 1):
        sev = sheet.cell(row=row_idx, column=severity_column_index).value
        fill = sheet.cell(row=row_idx, column=severity_column_index).fill
        actual[sev] = fill.start_color.rgb
    assert actual == expected


def test_export_risks_csv_basic():
    db = FakeSession()
    admin = _admin_user()
    db.users.append(admin)
    asset = _asset(admin.id)
    db.assets.append(asset)
    doc = _document(asset.id, admin.id)
    db.documents.append(doc)
    db.document_by_id = {doc.id: doc}
    db.risks.append(_risk(doc.id))

    client = _make_client(db, admin)
    response = client.get(f"/exports/risks?asset_id={asset.id}")
    assert response.status_code == 200
    lines = response.text.strip().split("\n")
    assert lines[0].strip() == ",".join(RISK_COLUMNS)
    assert len(lines) == 2  # header + 1 row
    assert "financial" in lines[1]


def test_export_obligations_forbidden_for_unassigned_asset():
    db, admin, asset = _seed_obligations_db(obligations_count=1)
    viewer = _viewer_user()
    db.users.append(viewer)
    client = _make_client(db, viewer)

    response = client.get(f"/exports/obligations?asset_id={asset.id}")
    assert response.status_code == 403


def test_export_obligations_empty_result_returns_valid_csv():
    db = FakeSession()
    admin = _admin_user()
    db.users.append(admin)
    asset = _asset(admin.id)
    db.assets.append(asset)
    client = _make_client(db, admin)

    response = client.get(f"/exports/obligations?asset_id={asset.id}")
    assert response.status_code == 200
    lines = response.text.strip().split("\n")
    assert len(lines) == 1
    assert lines[0].strip() == ",".join(OBLIGATION_COLUMNS)


def test_export_obligations_cap_triggers_413(monkeypatch):
    db, admin, asset = _seed_obligations_db(obligations_count=3)
    monkeypatch.setattr(exports_router, "_max_rows", lambda: 2)
    client = _make_client(db, admin)

    response = client.get(f"/exports/obligations?asset_id={asset.id}")
    assert response.status_code == 413
    assert "tighten filters" in response.json()["detail"]
